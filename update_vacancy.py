# -*- coding: utf-8 -*-
"""
ときわ苑ショートステイ予定表 Excel監視・自動解析スクリプト
同フォルダ内の「ときわ苑入所者一覧表.xlsx」を監視し、更新を検知したら
当月と翌月の通常枠の空き状況を解析して data.json へ書き出します。
"""

import os
import time
import json
import datetime
import traceback
import shutil

EXCEL_FILE = "C:/Users/tokiw/OneDrive/デスクトップ/ときわ苑入所者一覧表.xlsx"
JSON_FILE = "data.json"

def install_openpyxl():
    """openpyxlを自動インストールする"""
    import subprocess
    import sys
    print("openpyxl が見つかりません。自動インストールを実行します...")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        print("openpyxl のインストールが完了しました。")
        return True
    except Exception as e:
        print(f"openpyxlの自動インストールに失敗しました: {e}")
        print("手動で 'pip install openpyxl' を実行してください。")
        return False

# openpyxlのインポート確認と自動インストール
try:
    import openpyxl
except ImportError:
    if not install_openpyxl():
        sys.exit(1)
    import openpyxl

def parse_excel():
    """Excelファイルを解析して空き状況データをJSONに書き出す"""
    if not os.path.exists(EXCEL_FILE):
        print(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 警告: {EXCEL_FILE} が見つかりません。ファイルを配置してください。")
        return False
    
    print(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {EXCEL_FILE} を解析中...")

    try:
        # ExcelがExcel自身に開かれてロックされている可能性があるため、一時ファイルにコピーして読み込む
        temp_excel = "temp_load.xlsx"
        try:
            shutil.copy2(EXCEL_FILE, temp_excel)
            wb = openpyxl.load_workbook(temp_excel, data_only=True)
            # 読み込み終わったら一時ファイルを削除
            if os.path.exists(temp_excel):
                os.remove(temp_excel)
        except Exception as copy_err:
            print(f"一時ファイルでの読み込み失敗(直接読み込みを試行します): {copy_err}")
            wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

        today = datetime.date.today()
        
        # 当月と翌月の判定
        current_month = today.month
        next_month = (today.month % 12) + 1
        
        current_sheet_name = f"{current_month}月"
        next_sheet_name = f"{next_month}月"
        
        target_sheets = []
        for name in wb.sheetnames:
            if current_sheet_name in name:
                target_sheets.append(wb[name])
            elif next_sheet_name in name:
                target_sheets.append(wb[name])
                
        if not target_sheets:
            print(f"シート名に「{current_sheet_name}」または「{next_sheet_name}」を含むシートが見つかりませんでした。")
            return False
            
        # 既存 of the JSONデータをロードしてマージするための準備（別の日付データを上書きしないようにする）
        availability_data = {}
        if os.path.exists(JSON_FILE):
            try:
                with open(JSON_FILE, "r", encoding="utf-8") as f:
                    availability_data = json.load(f)
            except Exception:
                pass
                
        updated_count = 0
        hospital_limit_val = 3 # デフォルトの追加受入可能数

        # 1. 「短期原本」シートの BE1 セル (1行目 57列目) から直接数値を読み取る
        if "短期原本" in wb.sheetnames:
            orig_sheet = wb["短期原本"]
            be1_val = orig_sheet.cell(row=1, column=57).value
            if be1_val is not None:
                try:
                    hospital_limit_val = int(be1_val)
                    print(f"「短期原本」シートの BE1 セルから入院枠上限を検出: {hospital_limit_val}名")
                except ValueError:
                    pass

        for sheet in target_sheets:
            # 短期原本に数値がセットされていなかった場合のみ、各月シートの BE1 からフォールバック取得
            if hospital_limit_val == 3:
                be1_val = sheet.cell(row=1, column=57).value
                if be1_val is not None:
                    try:
                        hospital_limit_val = int(be1_val)
                        print(f"シート [{sheet.title}] の BE1 セルから入院枠上限を検出: {hospital_limit_val}名")
                    except ValueError:
                        pass

            # 2. 空き状況データの解析
            date_row = 3  # 3行目 (openpyxlは1始まり)
            date_columns = {}  # 列番号(1始まり) -> YYYY-MM-DD
            
            # 日付行をスキャンして日付列の位置を特定
            for col in range(2, sheet.max_column + 1):  # B列(2列目)以降
                val = sheet.cell(row=date_row, column=col).value
                if val:
                    date_str = ""
                    if isinstance(val, (datetime.date, datetime.datetime)):
                        date_str = val.strftime("%Y-%m-%d")
                    elif isinstance(val, str):
                        val_str = val.strip()
                        # YYYY-MM-DD
                        if len(val_str) >= 10 and val_str[4] == '-' and val_str[7] == '-':
                            date_str = val_str[:10]
                        # YYYY/MM/DD のスラッシュ対応
                        elif len(val_str) >= 10 and val_str[4] == '/' and val_str[7] == '/':
                            date_str = val_str[:10].replace('/', '-')
                    
                    if date_str:
                        date_columns[col] = date_str
            
            if not date_columns:
                continue
                
            # A列、B列、C列のいずれかが「利用可能床数(残)」の行を探索
            target_row_idx = -1
            for row in range(1, min(sheet.max_row + 1, 100)):
                for col in (1, 2, 3):
                    cell_val = sheet.cell(row=row, column=col).value
                    if cell_val and "利用可能" in str(cell_val) and "残" in str(cell_val):
                        target_row_idx = row
                        break
                if target_row_idx != -1:
                    break
                    
            if target_row_idx == -1:
                continue
                
            # 特定した日付列の値を抽出
            for col, date_str in date_columns.items():
                val = sheet.cell(row=target_row_idx, column=col).value
                if val is not None:
                    try:
                        num = int(val)
                        vacancy = num  # マイナス値（オーバー人数）もそのまま保持してJS側で判定する
                        
                        if date_str not in availability_data:
                            availability_data[date_str] = {}
                        availability_data[date_str]["regular"] = vacancy
                        updated_count += 1
                    except (ValueError, TypeError):
                        pass
                        
        if updated_count > 0:
            # JSONに保存 (予備)
            with open(JSON_FILE, "w", encoding="utf-8") as f:
                json.dump(availability_data, f, ensure_ascii=False, indent=4)
            
            # HTMLファイルに直接データを埋め込む (CORSエラー対策)
            update_html(availability_data, hospital_limit_val)
            
            print(f"成功: {updated_count}件の日付データと、入院枠上限: {hospital_limit_val}名を同期しました。")

            # Git自動同期処理
            import subprocess
            try:
                current_dir = os.path.dirname(os.path.abspath(__file__))
                status = subprocess.run(["git", "status", "--porcelain", "index.html"], capture_output=True, text=True, cwd=current_dir)
                if status.stdout.strip():
                    print("カレンダーの更新を検知しました。GitHubへ全自動同期します...")
                    subprocess.run(["git", "add", "index.html"], check=True, cwd=current_dir)
                    subprocess.run(["git", "commit", "-m", "Auto-update shortstay vacancy calendar"], check=True, cwd=current_dir)
                    subprocess.run(["git", "push"], check=True, cwd=current_dir)
                    print("GitHubへの同期に成功しました！Vercelが数秒後に自動でカレンダーを更新します。")
            except Exception as e:
                print(f"（Git自動同期スキップ: {e}）")

            return True
        else:
            print("警告: 該当行から有効な空き状況データを抽出できませんでした。")
            return False
            
    except Exception as e:
        print(f"エラー: Excelの解析に失敗しました: {e}")
        traceback.print_exc()
        return False

def update_html(availability_data, hospital_limit):
    """index.html の AVAILABILITY_DATA および DEFAULT_HOSPITAL_LIMIT を直接書き換える"""
    html_file = "C:/Users/tokiw/projects/shortstay-vacancy/index.html"
    if not os.path.exists(html_file):
        html_file = "index.html"
        if not os.path.exists(html_file):
            print("警告: index.html が見つかりません。")
            return False
        
    try:
        print(f"HTMLへ反映する入院枠上限: {hospital_limit}名")

        with open(html_file, "r", encoding="utf-8") as f:
            content = f.read()
            
        start_marker = "// === PYTHON_AUTO_REPLACE_START ==="
        end_marker = "// === PYTHON_AUTO_REPLACE_END ==="
        
        start_idx = content.find(start_marker)
        end_idx = content.find(end_marker)
        
        if start_idx == -1 or end_idx == -1:
            print("HTMLファイル内にデータ書き換え用のマーカーが見つかりませんでした。")
            return False
            
        # JSON 文字列を生成
        json_str = json.dumps(availability_data, ensure_ascii=False, indent=8)
        
        # 置換テキストの構築
        replacement = f"{start_marker}\n    let AVAILABILITY_DATA = {json_str};\n    "
        
        # AVAILABILITY_DATAの置換実行
        new_content = content[:start_idx] + replacement + content[end_idx:]
        
        # DEFAULT_HOSPITAL_LIMIT 定数の書き換え
        import re
        new_content = re.sub(
            r'const DEFAULT_HOSPITAL_LIMIT = \d+;',
            f'const DEFAULT_HOSPITAL_LIMIT = {hospital_limit};',
            new_content
        )
        
        with open(html_file, "w", encoding="utf-8") as f:
            f.write(new_content)
            
        print("index.html への直接書き込みに成功しました。")
        return True
    except Exception as e:
        print(f"HTML書き込みエラー: {e}")
        return False

def main():
    print("====================================================")
    print(" ときわ苑ショートステイ空き状況 自動監視スクリプト")
    print("====================================================")
    print(f"監視ファイル : {EXCEL_FILE}")
    print(f"出力JSON     : {JSON_FILE}")
    print("----------------------------------------------------")
    
    last_mtime = 0
    if os.path.exists(EXCEL_FILE):
        last_mtime = os.path.getmtime(EXCEL_FILE)
        print("初回解析を実行します...")
        parse_excel()
    else:
        print(f"※現在フォルダ内に「{EXCEL_FILE}」がありません。")
        print("  ファイルを配置すると自動検知して解析を開始します。")
        
    print("ファイルの変更監視を開始しました。この画面を開いたままにしてください。")
    print("(Ctrl+C キーで監視を終了します)\n")
    
    try:
        while True:
            if os.path.exists(EXCEL_FILE):
                current_mtime = os.path.getmtime(EXCEL_FILE)
                if current_mtime != last_mtime:
                    print(f"\nファイルの変更を検知しました: {EXCEL_FILE}")
                    # 解析実行。Excel保存中の書き込みロック一時回避のために少し待つ
                    time.sleep(0.5)
                    if parse_excel():
                        last_mtime = current_mtime
                    else:
                        print("解析に失敗したため、ファイル再更新時に再トライします。")
            time.sleep(2)
    except KeyboardInterrupt:
        print("\n監視プログラムを終了しました。")

if __name__ == "__main__":
    main()
